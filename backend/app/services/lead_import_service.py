import uuid
import io
import csv
import re
import base64
import httpx
import openpyxl
from datetime import datetime, timezone
from typing import List, Dict, Tuple, Any
from fastapi import HTTPException, status
from sqlalchemy.ext.asyncio import AsyncSession

import os
from app.models.user import User
from app.models.lead import Lead
from app.models.lead_import import LeadImport, LeadImportStatus
from app.core.redis import redis_client
from app.repositories.lead_import_repository import LeadImportRepository
from app.repositories.user_repository import UserRepository
from app.repositories.lead_repository import LeadRepository
from app.services.assignment_service import AssignmentService
from app.services.audit_service import AuditService
from app.services.import_.csv_parser import parse_csv
from app.services.import_.excel_parser import parse_excel
from app.services.import_.google_sheets_parser import fetch_google_sheet
from app.services.import_.validation_engine import validate_import_rows

# Define mappings and aliases for auto-matching
MAPPING_ALIASES = {
    "first_name": ["first name", "given name", "fname", "first"],
    "last_name": ["last name", "surname", "lname", "last"],
    "email": ["email", "email address", "mail", "email_address"],
    "phone": ["phone", "telephone", "phone number", "mobile", "phone_number"],
    "company_name": ["company", "organization", "org", "company name", "company_name"],
    "title": ["title", "job title", "position", "role", "job_title"],
    "value": ["value", "deal value", "deal_value", "opportunity value", "amount", "price", "deal amount", "deal_amount"],
    "source": ["source", "lead source", "lead_source", "channel"]
}

def calculate_mapping_confidence(header: str, field_name: str) -> float:
    header_clean = header.strip().lower().replace("_", " ").replace("-", " ")
    field_clean = field_name.strip().lower().replace("_", " ").replace("-", " ")
    
    if header_clean == field_clean:
        return 1.0
        
    aliases = MAPPING_ALIASES.get(field_name, [])
    for alias in aliases:
        alias_clean = alias.strip().lower().replace("_", " ").replace("-", " ")
        if header_clean == alias_clean:
            return 0.8
            
    return 0.0

def suggest_mappings(headers: List[str]) -> Dict[str, Dict[str, Any]]:
    suggestions = {}
    used_headers = set()
    fields = ["first_name", "last_name", "email", "phone", "company_name", "title", "value", "source"]
    
    for field in fields:
        best_header = None
        best_score = 0.0
        for header in headers:
            if header in used_headers:
                continue
            score = calculate_mapping_confidence(header, field)
            if score > best_score:
                best_score = score
                best_header = header
        if best_score > 0.0:
            suggestions[field] = {"column": best_header, "confidence": best_score}
            used_headers.add(best_header)
        else:
            suggestions[field] = {"column": None, "confidence": 0.0}
            
    return suggestions

class LeadImportService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.import_repo = LeadImportRepository(db)
        self.lead_repo = LeadRepository(db)
        self.user_repo = UserRepository(db)
        self.assignment_service = AssignmentService(db)
        self.audit_service = AuditService(db)

    @staticmethod
    def generate_csv_template() -> str:
        """Generate string of CSV template headers."""
        headers = ["First Name", "Last Name", "Email", "Phone", "Company", "Title", "Deal Value", "Source"]
        out = io.StringIO()
        writer = csv.writer(out)
        writer.writerow(headers)
        return out.getvalue()

    @staticmethod
    def generate_xlsx_template() -> bytes:
        """Generate bytes of XLSX template workbook."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leads Template"
        headers = ["First Name", "Last Name", "Email", "Phone", "Company", "Title", "Deal Value", "Source"]
        ws.append(headers)
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    @staticmethod
    def extract_google_sheets_csv_url(url: str) -> str:
        """Extract spreadsheet ID from URL and format as CSV export url."""
        match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
        if not match:
            raise ValueError("Invalid Google Sheets URL. Please check spreadsheet share link format.")
        spreadsheet_id = match.group(1)
        return f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=csv"

    @staticmethod
    def parse_csv_data(content: str) -> Tuple[List[str], List[Dict[str, Any]]]:
        """Parse raw CSV string into headers list and row dictionary list."""
        f = io.StringIO(content)
        reader = csv.reader(f)
        rows = list(reader)
        if not rows:
            return [], []
        # Filter empty rows
        rows = [r for r in rows if any(col.strip() for col in r)]
        if not rows:
            return [], []
        headers = [h.strip() for h in rows[0]]
        data_rows = []
        for r in rows[1:]:
            row_dict = {}
            for idx, col in enumerate(r):
                if idx < len(headers):
                    row_dict[headers[idx]] = col.strip()
            data_rows.append(row_dict)
        return headers, data_rows

    @staticmethod
    def parse_xlsx_data(file_bytes: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
        """Parse Excel bytes file into headers list and row dictionary list."""
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return [], []
        rows = [r for r in rows if any(col is not None for col in r)]
        if not rows:
            return [], []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        data_rows = []
        for r in rows[1:]:
            row_dict = {}
            for idx, col in enumerate(r):
                if idx < len(headers):
                    val = str(col).strip() if col is not None else ""
                    row_dict[headers[idx]] = val
            data_rows.append(row_dict)
        return headers, data_rows

    async def get_preview_from_file(self, filename: str, content: bytes) -> Dict[str, Any]:
        """Parse file content, cache data in Redis, and return headers, suggestions, and preview rows."""
        file_token = str(uuid.uuid4())
        
        try:
            if filename.endswith(".xlsx"):
                headers, rows = parse_excel(content)
                # Store in base64 format inside redis
                b64_content = base64.b64encode(content).decode("utf-8")
                await redis_client.set(f"import_file:{file_token}", b64_content, ex=3600)
            else:
                # Assume CSV
                headers, rows = parse_csv(content)
                text_content = content.decode("utf-8", errors="ignore")
                await redis_client.set(f"import_file:{file_token}", text_content, ex=3600)
        except ValueError as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=str(e)
            )
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Failed to parse file: {str(e)}"
            )

        if not headers:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Empty file or missing headers row"
            )

        suggested_mapping = suggest_mappings(headers)
        preview_rows = rows[:10]

        return {
            "file_token": file_token,
            "headers": headers,
            "suggested_mapping": suggested_mapping,
            "preview_rows": preview_rows
        }

    async def get_preview_from_google_sheets(self, sheet_url: str) -> Dict[str, Any]:
        """Fetch Google Sheets publicly shared link, parse, cache in Redis, and return preview."""
        try:
            csv_content = await fetch_google_sheet(sheet_url)
        except ValueError as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=str(e)
            )
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Google Sheets fetch error: {str(e)}"
            )

        file_token = str(uuid.uuid4())
        await redis_client.set(f"import_file:{file_token}", csv_content, ex=3600)

        try:
            headers, rows = parse_csv(csv_content.encode("utf-8"))
        except ValueError as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=str(e)
            )

        if not headers:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Google Sheet does not contain columns or headers"
            )

        suggested_mapping = suggest_mappings(headers)
        preview_rows = rows[:10]

        return {
            "file_token": file_token,
            "headers": headers,
            "suggested_mapping": suggested_mapping,
            "preview_rows": preview_rows
        }

    async def process_import_batch(
        self, 
        actor: User, 
        file_token: str, 
        source_type: str, 
        column_mapping: Dict[str, str], 
        auto_assign: bool = True,
        assignment_mode: str = "NONE",
        assigned_user_id: uuid.UUID | None = None
    ) -> LeadImport:
        """
        Validate and import bulk leads under organization isolation.
        Saves successfully created leads and tracks/compiles failed rows report.
        """
        # Validate SPECIFIC_USER assignment target
        target_user = None
        if assignment_mode == "SPECIFIC_USER":
            if not assigned_user_id:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Assigned user ID is required when assignment mode is SPECIFIC_USER"
                )
            target_user = await self.user_repo.get_user_by_id(actor.organization_id, assigned_user_id)
            if not target_user or not target_user.is_active or target_user.role != "Employee":
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Invalid assignee. User must be an active employee in your organization."
                )

        raw_content = await redis_client.get(f"import_file:{file_token}")
        if not raw_content:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Import file cache expired. Please re-upload your file."
            )

        # Parse content
        try:
            if source_type == "file" and not raw_content.startswith("First Name") and not raw_content.startswith("Last Name") and "," not in raw_content[:100]:
                # Guess XLSX Base64 content
                try:
                    xlsx_bytes = base64.b64decode(raw_content)
                    headers, data_rows = parse_excel(xlsx_bytes)
                except Exception:
                    headers, data_rows = parse_csv(raw_content.encode("utf-8"))
            else:
                headers, data_rows = parse_csv(raw_content.encode("utf-8"))
        except ValueError as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=str(e)
            )

        # 1. Initialize Log record
        import_record = LeadImport(
            id=uuid.uuid4(),
            organization_id=actor.organization_id,
            filename="Uploaded_File.csv" if source_type == "file" else "Google_Sheet",
            status=LeadImportStatus.PROCESSING,
            total_rows=len(data_rows),
            successful_rows=0,
            failed_rows=0,
            mapping_confidence=0.0,
            error_summary=[],
            failed_rows_file_path=None,
            created_by=actor.id
        )
        self.db.add(import_record)
        await self.db.flush()

        # Calculate average mapping confidence score
        score_sum = 0.0
        mapped_count = 0
        for field, col in column_mapping.items():
            if col:
                score_sum += calculate_mapping_confidence(col, field)
                mapped_count += 1
        avg_confidence = score_sum / mapped_count if mapped_count > 0 else 0.0
        import_record.mapping_confidence = avg_confidence

        # Define email check function for validation engine
        async def check_existing_email(email: str) -> bool:
            dup = await self.lead_repo.get_lead_by_email(actor.organization_id, email)
            return dup is not None

        # 2. Use Validation Engine to validate all rows
        valid_rows, errors_log = await validate_import_rows(
            rows=data_rows,
            column_mapping=column_mapping,
            check_existing_email_fn=check_existing_email
        )

        success_count = 0
        fail_count = len(errors_log)

        # Keep track of failed rows structure for CSV report
        failed_csv_out = io.StringIO()
        csv_writer = csv.writer(failed_csv_out)
        csv_writer.writerow(headers + ["Import Error Reason"])

        # Map rows in error log to actual CSV rows and append to CSV writer
        errors_by_row = {err["row"]: err["reason"] for err in errors_log}
        for idx, row in enumerate(data_rows, start=2):
            if idx in errors_by_row:
                raw_row_vals = [row.get(h, "") for h in headers]
                csv_writer.writerow(raw_row_vals + [errors_by_row[idx]])

        # Insert valid rows
        for item in valid_rows:
            mapped_values = item["data"]
            row_idx = item["row_index"] - 2
            raw_row = data_rows[row_idx]
            
            # Extract opportunity value
            val_col = column_mapping.get("value")
            val_str = raw_row.get(val_col, "").strip() if val_col else ""
            val = None
            if val_str:
                try:
                    val = float(val_str.replace("$", "").replace(",", "").strip())
                except ValueError:
                    pass

            lead_obj = Lead(
                organization_id=actor.organization_id,
                first_name=mapped_values.get("first_name") or None,
                last_name=mapped_values.get("last_name"),
                email=mapped_values.get("email") or None,
                phone=mapped_values.get("phone") or None,
                company_name=mapped_values.get("company_name") or None,
                title=mapped_values.get("title"),
                status=mapped_values.get("status") or "New",
                source=mapped_values.get("source") or "Import",
                value=val,
                created_by=actor.id,
                import_id=import_record.id
            )
            self.db.add(lead_obj)
            await self.db.flush()

            if assignment_mode == "AUTO":
                await self.assignment_service.assign_lead(lead_obj)
            elif assignment_mode == "SPECIFIC_USER" and target_user:
                await self.assignment_service.assign_lead_to_user(lead_obj, target_user)
            
            success_count += 1

        # Commit transaction on successfully processed leads
        import_record.successful_rows = success_count
        import_record.failed_rows = fail_count
        import_record.error_summary = errors_log
        
        # Save failed rows report to filesystem
        if fail_count > 0:
            uploads_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "uploads")
            os.makedirs(uploads_dir, exist_ok=True)
            file_path = os.path.join(uploads_dir, f"failed_rows_{import_record.id}.csv")
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(failed_csv_out.getvalue())
            import_record.failed_rows_file_path = file_path
        else:
            import_record.failed_rows_file_path = None
            
        import_record.status = LeadImportStatus.COMPLETED

        self.db.add(import_record)
        await self.db.flush()

        # Log audit entry
        await self.audit_service.log_event(
            organization_id=actor.organization_id,
            actor_user_id=actor.id,
            action="LEAD_IMPORT_COMPLETED",
            resource_type="import",
            resource_id=str(import_record.id),
            action_metadata={
                "total_rows": len(data_rows),
                "success_count": success_count,
                "fail_count": fail_count,
                "assignment_mode": assignment_mode,
                "assigned_user_id": str(assigned_user_id) if assigned_user_id else None
            }
        )

        # Clear file cache from Redis
        await redis_client.delete(f"import_file:{file_token}")

        return import_record

    async def get_failed_rows_report(self, organization_id: uuid.UUID, import_id: uuid.UUID) -> str:
        """Retrieve CSV report text of validation failures for downloading."""
        import_job = await self.import_repo.get_import_by_id(organization_id, import_id)
        if not import_job or not import_job.failed_rows_file_path:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Failed rows report not found for this import"
            )
        try:
            with open(import_job.failed_rows_file_path, "r", encoding="utf-8") as f:
                return f.read()
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Failed to read report file: {str(e)}"
            )
