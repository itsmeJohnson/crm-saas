import io
import csv
import openpyxl
from openpyxl.utils import get_column_letter

TEMPLATE_COLUMNS = [
    "first_name",
    "last_name",
    "email",
    "phone",
    "company_name",
    "title",
    "source",
    "status",
    "assigned_email"
]

def generate_csv_template() -> str:
    """Generate standardized CSV template header string."""
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(TEMPLATE_COLUMNS)
    return out.getvalue()

def generate_xlsx_template() -> bytes:
    """Generate standardized Excel (.xlsx) template binary content."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads Template"
    
    ws.append(TEMPLATE_COLUMNS)
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
