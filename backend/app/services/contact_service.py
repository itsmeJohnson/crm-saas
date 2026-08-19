import uuid
from datetime import datetime, timezone
from typing import Sequence, Tuple
from fastapi import HTTPException, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from app.repositories.contact_repository import ContactRepository
from app.repositories.company_repository import CompanyRepository
from app.repositories.user_repository import UserRepository
from app.repositories.activity_repository import ActivityRepository
from app.repositories.note_repository import NoteRepository
from app.services.audit_service import AuditService
from app.services.dashboard_service import DashboardService
from app.services.notification_service import NotificationService
from app.models.user import User
from app.models.contact import Contact

class ContactService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.contact_repo = ContactRepository(db)
        self.company_repo = CompanyRepository(db)
        self.user_repo = UserRepository(db)
        self.activity_repo = ActivityRepository(db)
        self.note_repo = NoteRepository(db)
        self.audit_service = AuditService(db)
        self.notification_service = NotificationService(db)

    async def _apply_custom_fields(
        self,
        actor: User,
        contact_data: dict,
        existing_contact: Contact | None = None,
    ) -> None:
        """Validate + sanitize contact custom fields with the SAME type-aware
        MetadataValidationEngine used for Leads (G2), and write the sanitized
        result back into ``contact_data['custom_fields']``.

        On update, merges the incoming payload over the stored values so partial
        updates keep untouched fields, and drops keys explicitly cleared with an
        empty/None value (mirrors the Lead update semantics).
        """
        if "custom_fields" not in contact_data:
            return
        from app.services.custom_field_service import CustomFieldService
        from app.services.metadata_validation_engine import (
            MetadataValidationEngine,
            MetadataValidationError,
        )

        cf_service = CustomFieldService(self.db)
        definitions = await cf_service.list_definitions(actor, "contact")
        incoming_cf = contact_data.get("custom_fields") or {}

        if existing_contact is not None:
            def_map = {d.key: d for d in definitions if d.is_active}
            merged = dict(existing_contact.custom_fields or {})
            for key, val in incoming_cf.items():
                definition = def_map.get(key)
                is_required = bool(definition and (definition.validation_rules or {}).get("required") is True)
                if (val is None or val == "") and not is_required:
                    merged.pop(key, None)
                else:
                    merged[key] = val
            payload = merged
            exclude_id = existing_contact.id
        else:
            payload = incoming_cf
            exclude_id = None

        try:
            sanitized = await MetadataValidationEngine.validate_and_sanitize(
                self.db, Contact, actor.organization_id, definitions, payload, exclude_id=exclude_id
            )
        except MetadataValidationError as e:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
        contact_data["custom_fields"] = sanitized

    async def _notify_assignment(self, actor: User, contact: Contact) -> None:
        if contact.assigned_user_id and contact.assigned_user_id != actor.id:
            await self.notification_service.create_notification(
                organization_id=actor.organization_id,
                user_id=contact.assigned_user_id,
                category="contact",
                title="Contact assigned to you",
                body=f"{contact.first_name} {contact.last_name} was assigned to you.",
                link_url=f"/contacts?contactId={contact.id}",
                action_metadata={"contact_id": str(contact.id)},
            )

    async def get_contact(self, actor: User, contact_id: uuid.UUID) -> Contact:
        if not actor.is_active:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Actor is inactive")
        
        contact = await self.contact_repo.get_contact_by_id(actor.organization_id, contact_id)
        if not contact:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Contact not found")
        return contact

    async def create_contact(self, actor: User, contact_data: dict) -> Contact:
        if not actor.is_active:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Actor is inactive")

        # Block duplicate patients (same phone/email) unless explicitly overridden.
        # `allow_duplicate` is a control flag, not a column — pop before persisting.
        allow_duplicate = contact_data.pop("allow_duplicate", False)
        if not allow_duplicate and (contact_data.get("email") or contact_data.get("phone")):
            dupes = await self.contact_repo.find_duplicates(
                actor.organization_id,
                email=contact_data.get("email"),
                phone=contact_data.get("phone"),
            )
            if dupes:
                existing = dupes[0]
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail={
                        "message": (f"A contact with this phone/email already exists: "
                                    f"{existing.first_name} {existing.last_name}. "
                                    f"Set allow_duplicate to create anyway."),
                        "existing_id": str(existing.id),
                    },
                )

        # Validate company reference
        company_id = contact_data.get("company_id")
        if company_id:
            company = await self.company_repo.get_company_by_id(actor.organization_id, company_id)
            if not company:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Company not found in your organization"
                )

        # Validate assigned user organization
        assigned_user_id = contact_data.get("assigned_user_id")
        if assigned_user_id:
            assigned_user = await self.user_repo.get_user_by_id(actor.organization_id, assigned_user_id)
            if not assigned_user or not assigned_user.is_active:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Assigned user not found or inactive in your organization"
                )

        await self._apply_custom_fields(actor, contact_data)

        contact = await self.contact_repo.create_contact(actor.organization_id, contact_data, actor.id)

        await self.audit_service.log_event(
            organization_id=actor.organization_id,
            actor_user_id=actor.id,
            action="CONTACT_CREATED",
            resource_type="contact",
            resource_id=str(contact.id),
            action_metadata={"email": contact.email, "name": f"{contact.first_name} {contact.last_name}"}
        )
        await self._notify_assignment(actor, contact)

        # Run contact-created automation rules
        from app.services.workflow_service import WorkflowService
        await WorkflowService(self.db).run("contact_created", contact, actor, entity_type="contact")

        await DashboardService.invalidate_cache(actor.organization_id)
        return await self.contact_repo.get_contact_by_id(actor.organization_id, contact.id)

    async def paginate_contacts(
        self,
        actor: User,
        skip: int = 0,
        limit: int = 100,
        search_query: str | None = None,
        company_id: uuid.UUID | None = None,
        assigned_user_id: uuid.UUID | None = None,
        tag: str | None = None,
        has_email: bool | None = None,
        created_from=None,
        created_to=None,
    ) -> Tuple[Sequence[Contact], int]:
        if not actor.is_active:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Actor is inactive")

        if company_id:
            # Verify company belongs to tenant
            company = await self.company_repo.get_company_by_id(actor.organization_id, company_id)
            if not company:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Company not found in your organization"
                )

        return await self.contact_repo.paginate_contacts(
            actor.organization_id, skip, limit, search_query, company_id,
            assigned_user_id=assigned_user_id, tag=tag, has_email=has_email,
            created_from=created_from, created_to=created_to,
        )

    async def update_contact(self, actor: User, contact_id: uuid.UUID, contact_data: dict) -> Contact:
        if not actor.is_active:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Actor is inactive")

        contact = await self.get_contact(actor, contact_id)

        # Field-level permission check (no-op unless actor has a custom role)
        from app.services.permission_service import PermissionService
        await PermissionService(self.db).enforce_field_writes(actor, "contacts", contact_data)

        # Validate company reference if updated
        company_id = contact_data.get("company_id")
        if company_id:
            company = await self.company_repo.get_company_by_id(actor.organization_id, company_id)
            if not company:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Company not found in your organization"
                )

        # Validate assigned user organization if updated
        assigned_user_id = contact_data.get("assigned_user_id")
        if assigned_user_id:
            assigned_user = await self.user_repo.get_user_by_id(actor.organization_id, assigned_user_id)
            if not assigned_user or not assigned_user.is_active:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Assigned user not found or inactive in your organization"
                )

        await self._apply_custom_fields(actor, contact_data, existing_contact=contact)

        prev_assignee = contact.assigned_user_id
        updated = await self.contact_repo.update_contact(actor.organization_id, contact_id, contact_data)

        await self.audit_service.log_event(
            organization_id=actor.organization_id,
            actor_user_id=actor.id,
            action="CONTACT_UPDATED",
            resource_type="contact",
            resource_id=str(contact_id),
            action_metadata={"updated_fields": list(contact_data.keys())}
        )

        # Notify only on a genuine assignment change
        if "assigned_user_id" in contact_data and updated.assigned_user_id != prev_assignee:
            await self._notify_assignment(actor, updated)

        from app.services.workflow_service import WorkflowService
        await WorkflowService(self.db).run("contact_updated", updated, actor, entity_type="contact")

        await DashboardService.invalidate_cache(actor.organization_id)
        return await self.contact_repo.get_contact_by_id(actor.organization_id, contact_id)

    async def soft_delete_contact(self, actor: User, contact_id: uuid.UUID) -> Contact:
        if not actor.is_active:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Actor is inactive")

        contact = await self.get_contact(actor, contact_id)

        deleted = await self.contact_repo.soft_delete_contact(actor.organization_id, contact_id)

        # Cascade soft-delete activities and notes
        await self.activity_repo.soft_delete_by_parent(actor.organization_id, "contact", contact_id)
        await self.note_repo.soft_delete_by_parent(actor.organization_id, "contact", contact_id)

        await self.audit_service.log_event(
            organization_id=actor.organization_id,
            actor_user_id=actor.id,
            action="CONTACT_DELETED",
            resource_type="contact",
            resource_id=str(contact_id)
        )
        await DashboardService.invalidate_cache(actor.organization_id)
        return deleted

    # --- Duplicate detection ---
    async def find_duplicates(self, actor: User, email: str | None = None, phone: str | None = None,
                              exclude_contact_id: uuid.UUID | None = None) -> Sequence[Contact]:
        if not actor.is_active:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Actor is inactive")
        if not email and not phone:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Provide an email or phone to search")
        return await self.contact_repo.find_duplicates(actor.organization_id, email=email, phone=phone, exclude_contact_id=exclude_contact_id)

    # --- Export ---
    EXPORT_COLUMNS = ["first_name", "last_name", "email", "phone", "job_title", "company", "tags", "created_at"]

    async def export_contacts(self, actor: User, filters: dict) -> list[dict]:
        if not actor.is_active:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Actor is inactive")
        contacts = await self.contact_repo.stream_for_export(actor.organization_id, **filters)
        # Resolve company names in one query
        company_ids = {c.company_id for c in contacts if c.company_id}
        names = {}
        if company_ids:
            from app.models.company import Company
            res = await self.db.execute(select(Company.id, Company.name).filter(Company.id.in_(company_ids)))
            names = {cid: cname for cid, cname in res.all()}
        rows = []
        for c in contacts:
            rows.append({
                "first_name": c.first_name or "", "last_name": c.last_name or "",
                "email": c.email or "", "phone": c.phone or "", "job_title": c.job_title or "",
                "company": names.get(c.company_id, "") if c.company_id else "",
                "tags": ", ".join(c.tags) if c.tags else "",
                "created_at": c.created_at.isoformat() if c.created_at else "",
            })
        return rows

    @staticmethod
    def build_export_csv(rows: list[dict]) -> str:
        import csv, io
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=ContactService.EXPORT_COLUMNS)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)
        return buf.getvalue()

    @staticmethod
    def build_export_xlsx(rows: list[dict]) -> bytes:
        import io, openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contacts"
        ws.append(ContactService.EXPORT_COLUMNS)
        for r in rows:
            ws.append([r.get(col, "") for col in ContactService.EXPORT_COLUMNS])
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    # --- Import ---
    async def import_contacts(self, actor: User, content: bytes, filename: str) -> dict:
        """Parse a CSV/XLSX file and create contacts. Returns {created, failed, errors}."""
        if not actor.is_active:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Actor is inactive")
        rows = self._parse_import_file(content, filename)
        created, failed, errors = 0, 0, []

        # Preload companies for name->id matching (and create missing)
        from app.models.company import Company
        comp_res = await self.db.execute(
            select(Company).filter(Company.organization_id == actor.organization_id, Company.is_deleted == False)
        )
        companies = {c.name.lower(): c for c in comp_res.scalars().all()}

        for idx, row in enumerate(rows, start=2):
            g = lambda *keys: next((row[k].strip() for k in keys if k in row and row[k] and row[k].strip()), "")
            first = g("first_name", "first name", "firstname")
            last = g("last_name", "last name", "lastname")
            if not first or not last:
                failed += 1
                errors.append({"row": idx, "reason": "first_name and last_name are required"})
                continue
            company_name = g("company", "company_name", "company name")
            company_id = None
            if company_name:
                comp = companies.get(company_name.lower())
                if not comp:
                    comp = Company(organization_id=actor.organization_id, name=company_name, created_by=actor.id)
                    self.db.add(comp)
                    await self.db.flush()
                    companies[company_name.lower()] = comp
                company_id = comp.id
            contact = Contact(
                organization_id=actor.organization_id,
                first_name=first[:100], last_name=last[:100],
                email=(g("email") or None), phone=(g("phone") or None),
                job_title=(g("job_title", "job title", "title") or None),
                company_id=company_id, created_by=actor.id,
            )
            self.db.add(contact)
            created += 1

        await self.db.flush()
        await self.audit_service.log_event(
            organization_id=actor.organization_id, actor_user_id=actor.id,
            action="CONTACT_IMPORTED", resource_type="contact", resource_id=None,
            action_metadata={"created": created, "failed": failed},
        )
        await DashboardService.invalidate_cache(actor.organization_id)
        return {"created": created, "failed": failed, "errors": errors}

    @staticmethod
    def _parse_import_file(content: bytes, filename: str) -> list[dict]:
        import csv, io, os
        ext = os.path.splitext(filename)[1].lower().lstrip(".")
        if ext in ("xlsx", "xls"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            try:
                headers = [str(h).strip().lower() if h is not None else "" for h in next(rows_iter)]
            except StopIteration:
                return []
            parsed = []
            for r in rows_iter:
                parsed.append({headers[i]: ("" if v is None else str(v)) for i, v in enumerate(r) if i < len(headers)})
            return parsed
        else:
            try:
                text = content.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = content.decode("latin-1")
            reader = csv.DictReader(io.StringIO(text))
            return [{(k or "").strip().lower(): (v or "") for k, v in r.items()} for r in reader]

    # --- Bulk actions ---
    async def bulk_update(self, actor: User, contact_ids: list[uuid.UUID], fields: dict) -> dict:
        if not actor.is_active:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Actor is inactive")
        if fields.get("company_id"):
            company = await self.company_repo.get_company_by_id(actor.organization_id, fields["company_id"])
            if not company:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Company not found in your organization")
        if fields.get("assigned_user_id"):
            u = await self.user_repo.get_user_by_id(actor.organization_id, fields["assigned_user_id"])
            if not u or not u.is_active:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Assigned user not found or inactive")

        contacts = await self.contact_repo.get_contacts_for_update(actor.organization_id, contact_ids)
        add_tags = fields.get("add_tags") or []
        remove_tags = set(fields.get("remove_tags") or [])
        affected = []
        for c in contacts:
            if "company_id" in fields and fields["company_id"] is not None:
                c.company_id = fields["company_id"]
            if "assigned_user_id" in fields and fields["assigned_user_id"] is not None:
                c.assigned_user_id = fields["assigned_user_id"]
            if add_tags or remove_tags:
                current = list(c.tags or [])
                for t in add_tags:
                    if t not in current:
                        current.append(t)
                current = [t for t in current if t not in remove_tags]
                c.tags = current
            self.db.add(c)
            affected.append(c.id)
        await self.db.flush()
        if affected:
            await self.audit_service.log_event(
                organization_id=actor.organization_id, actor_user_id=actor.id,
                action="CONTACT_BULK_UPDATED", resource_type="contact", resource_id=None,
                action_metadata={"count": len(affected), "fields": list(fields.keys())},
            )
            await DashboardService.invalidate_cache(actor.organization_id)
        return {"affected_count": len(affected), "contact_ids": affected}

    async def bulk_delete(self, actor: User, contact_ids: list[uuid.UUID]) -> dict:
        if not actor.is_active:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Actor is inactive")
        contacts = await self.contact_repo.get_contacts_for_update(actor.organization_id, contact_ids)
        affected = []
        for c in contacts:
            c.is_deleted = True
            c.deleted_at = datetime.now(timezone.utc)
            self.db.add(c)
            await self.activity_repo.soft_delete_by_parent(actor.organization_id, "contact", c.id)
            await self.note_repo.soft_delete_by_parent(actor.organization_id, "contact", c.id)
            affected.append(c.id)
        await self.db.flush()
        if affected:
            await self.audit_service.log_event(
                organization_id=actor.organization_id, actor_user_id=actor.id,
                action="CONTACT_BULK_DELETED", resource_type="contact", resource_id=None,
                action_metadata={"count": len(affected)},
            )
            await DashboardService.invalidate_cache(actor.organization_id)
        return {"affected_count": len(affected), "contact_ids": affected}

    # --- Attachments ---
    ATTACHMENT_EXTENSIONS = {"pdf", "png", "jpg", "jpeg", "webp", "csv", "xlsx", "docx"}

    async def add_attachment(self, actor: User, contact_id: uuid.UUID, content: bytes, filename: str) -> dict:
        from app.core.storage import validate_and_sanitize_file, get_storage_provider
        contact = await self.get_contact(actor, contact_id)
        try:
            sanitized, ext = validate_and_sanitize_file(content=content, filename=filename, allowed_extensions=self.ATTACHMENT_EXTENSIONS)
        except ValueError as e:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
        url = await get_storage_provider().upload_file(content, sanitized)
        attachment = {
            "filename": filename, "stored_name": sanitized, "url": url, "size": len(content),
            "uploaded_by": str(actor.id), "uploaded_at": datetime.now(timezone.utc).isoformat(),
        }
        existing = list(contact.attachments or [])
        existing.append(attachment)
        contact.attachments = existing
        self.db.add(contact)
        await self.db.flush()
        await self.audit_service.log_event(
            organization_id=actor.organization_id, actor_user_id=actor.id,
            action="CONTACT_ATTACHMENT_ADDED", resource_type="contact", resource_id=str(contact_id),
            action_metadata={"filename": filename},
        )
        return attachment

    async def list_attachments(self, actor: User, contact_id: uuid.UUID) -> list[dict]:
        contact = await self.get_contact(actor, contact_id)
        return list(contact.attachments or [])

    async def delete_attachment(self, actor: User, contact_id: uuid.UUID, stored_name: str) -> dict:
        from app.core.storage import get_storage_provider
        contact = await self.get_contact(actor, contact_id)
        existing = list(contact.attachments or [])
        target = next((a for a in existing if a.get("stored_name") == stored_name or a.get("filename") == stored_name), None)
        if not target:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Attachment not found")
        contact.attachments = [a for a in existing if a is not target]
        self.db.add(contact)
        await self.db.flush()
        try:
            await get_storage_provider().delete_file(target["url"])
        except Exception:
            pass
        return {"deleted": True}

    # --- Timeline / communications ---
    async def get_timeline(self, actor: User, contact_id: uuid.UUID) -> list[dict]:
        from sqlalchemy import or_
        from app.models.note import Note
        from app.models.activity import Activity
        from app.models.audit_log import AuditLog
        contact = await self.get_contact(actor, contact_id)
        events: list[dict] = []
        notes = await self.db.execute(select(Note).filter(Note.contact_id == contact.id, Note.is_deleted == False))
        for n in notes.scalars().all():
            events.append({"type": "note", "id": str(n.id), "timestamp": n.created_at, "title": "Note added",
                           "description": n.content, "actor_user_id": str(n.created_by) if n.created_by else None, "event_metadata": None})
        acts = await self.db.execute(select(Activity).filter(Activity.contact_id == contact.id, Activity.is_deleted == False))
        for a in acts.scalars().all():
            events.append({"type": "activity", "id": str(a.id), "timestamp": a.created_at,
                           "title": f"{a.activity_type}: {a.subject}", "description": a.description,
                           "actor_user_id": str(a.assigned_user_id) if a.assigned_user_id else (str(a.created_by) if a.created_by else None),
                           "event_metadata": {"status": a.status}})
        audits = await self.db.execute(select(AuditLog).filter(
            AuditLog.organization_id == actor.organization_id, AuditLog.resource_id == str(contact.id),
            or_(AuditLog.resource_type == "contact", AuditLog.resource_type == "Contact")))
        for al in audits.scalars().all():
            events.append({"type": "audit", "id": str(al.id), "timestamp": al.created_at, "title": al.action,
                           "description": None, "actor_user_id": str(al.actor_user_id) if al.actor_user_id else None,
                           "event_metadata": al.action_metadata})
        events.sort(key=lambda e: e["timestamp"], reverse=True)
        return events

    async def get_communications(self, actor: User, contact_id: uuid.UUID) -> list[dict]:
        from app.models.activity import Activity
        contact = await self.get_contact(actor, contact_id)
        res = await self.db.execute(
            select(Activity).filter(
                Activity.contact_id == contact.id,
                Activity.is_deleted == False,
                Activity.activity_type.in_(["Call", "Email"]),
            ).order_by(Activity.created_at.desc())
        )
        out = []
        for a in res.scalars().all():
            out.append({
                "id": str(a.id), "channel": a.activity_type, "subject": a.subject,
                "description": a.description, "direction": a.call_direction, "status": a.status,
                "timestamp": a.created_at, "recording_url": a.recording_url,
            })
        return out

    # --- Tags ---
    async def get_tags(self, actor: User) -> list[str]:
        res = await self.db.execute(
            select(Contact.tags).filter(
                Contact.organization_id == actor.organization_id,
                Contact.is_deleted == False,
                Contact.tags.isnot(None),
            )
        )
        seen = set()
        for (tags,) in res.all():
            for t in (tags or []):
                seen.add(t)
        return sorted(seen)

    # --- Report ---
    async def get_contact_report(self, actor: User, date_from=None, date_to=None) -> dict:
        from sqlalchemy import func
        from app.models.company import Company

        def base(*cols):
            q = select(*cols).select_from(Contact).filter(
                Contact.organization_id == actor.organization_id, Contact.is_deleted == False)
            if date_from is not None:
                q = q.filter(Contact.created_at >= date_from)
            if date_to is not None:
                q = q.filter(Contact.created_at <= date_to)
            return q

        total = (await self.db.execute(base(func.count(Contact.id)))).scalar() or 0
        with_email = (await self.db.execute(base(func.count(Contact.id)).filter(Contact.email.isnot(None), Contact.email != ""))).scalar() or 0
        with_phone = (await self.db.execute(base(func.count(Contact.id)).filter(Contact.phone.isnot(None), Contact.phone != ""))).scalar() or 0
        with_company = (await self.db.execute(base(func.count(Contact.id)).filter(Contact.company_id.isnot(None)))).scalar() or 0

        # by company
        comp_rows = (await self.db.execute(base(Contact.company_id, func.count(Contact.id)).group_by(Contact.company_id))).all()
        comp_ids = [r[0] for r in comp_rows if r[0]]
        comp_names = {}
        if comp_ids:
            cn = await self.db.execute(select(Company.id, Company.name).filter(Company.id.in_(comp_ids)))
            comp_names = {cid: name for cid, name in cn.all()}
        by_company = [{"label": comp_names.get(r[0], "No company") if r[0] else "No company", "count": r[1]} for r in comp_rows]

        # by owner
        from app.models.user import User as UserModel
        owner_rows = (await self.db.execute(base(Contact.assigned_user_id, func.count(Contact.id)).group_by(Contact.assigned_user_id))).all()
        owner_ids = [r[0] for r in owner_rows if r[0]]
        owner_names = {}
        if owner_ids:
            un = await self.db.execute(select(UserModel.id, UserModel.first_name, UserModel.last_name, UserModel.email).filter(UserModel.id.in_(owner_ids)))
            for uid, fn, ln, em in un.all():
                owner_names[uid] = f"{fn or ''} {ln or ''}".strip() or em
        by_owner = [{"label": owner_names.get(r[0], "Unassigned") if r[0] else "Unassigned", "count": r[1]} for r in owner_rows]

        # by tag (in python; tags is a JSON array)
        tag_rows = (await self.db.execute(base(Contact.tags).filter(Contact.tags.isnot(None)))).all()
        tag_counts: dict[str, int] = {}
        for (tags,) in tag_rows:
            for t in (tags or []):
                tag_counts[t] = tag_counts.get(t, 0) + 1
        by_tag = [{"label": k, "count": v} for k, v in sorted(tag_counts.items(), key=lambda x: -x[1])]

        return {
            "total_contacts": total, "with_email": with_email, "with_phone": with_phone,
            "with_company": with_company, "by_company": by_company, "by_owner": by_owner, "by_tag": by_tag,
        }

    # --- Merge (primary wins, fill blanks) ---
    async def merge_contacts(self, actor: User, primary_id: uuid.UUID, secondary_id: uuid.UUID) -> Contact:
        if primary_id == secondary_id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Cannot merge a contact with itself")
        primary = await self.get_contact(actor, primary_id)
        secondary = await self.get_contact(actor, secondary_id)

        # Fill blank scalar fields on primary from secondary
        for field in ("email", "phone", "job_title", "company_id", "assigned_user_id"):
            if not getattr(primary, field) and getattr(secondary, field):
                setattr(primary, field, getattr(secondary, field))

        # Union tags
        merged_tags = list(primary.tags or [])
        for t in (secondary.tags or []):
            if t not in merged_tags:
                merged_tags.append(t)
        primary.tags = merged_tags or None

        # Merge custom_fields (primary wins)
        merged_cf = dict(secondary.custom_fields or {})
        merged_cf.update(primary.custom_fields or {})
        primary.custom_fields = merged_cf or None

        # Concatenate attachments
        primary.attachments = (list(primary.attachments or []) + list(secondary.attachments or [])) or None

        # Re-point notes, activities, converted leads
        from sqlalchemy import or_ as _or
        from app.models.note import Note
        from app.models.activity import Activity
        from app.models.contact_relationship import ContactRelationship
        from app.models.lead import Lead
        for model in (Note, Activity):
            res = await self.db.execute(select(model).filter(model.contact_id == secondary_id))
            for obj in res.scalars().all():
                obj.contact_id = primary_id
                self.db.add(obj)
        lead_res = await self.db.execute(select(Lead).filter(Lead.converted_contact_id == secondary_id))
        for lead in lead_res.scalars().all():
            lead.converted_contact_id = primary_id
            self.db.add(lead)
        # Relationships: re-point both sides, drop self-references
        rel_res = await self.db.execute(select(ContactRelationship).filter(
            _or(ContactRelationship.contact_id == secondary_id, ContactRelationship.related_contact_id == secondary_id)
        ))
        for rel in rel_res.scalars().all():
            if rel.contact_id == secondary_id:
                rel.contact_id = primary_id
            if rel.related_contact_id == secondary_id:
                rel.related_contact_id = primary_id
            if rel.contact_id == rel.related_contact_id:
                rel.is_deleted = True
            self.db.add(rel)

        self.db.add(primary)

        # Soft-delete secondary
        secondary.is_deleted = True
        secondary.deleted_at = datetime.now(timezone.utc)
        self.db.add(secondary)
        await self.db.flush()

        await self.audit_service.log_event(
            organization_id=actor.organization_id, actor_user_id=actor.id,
            action="CONTACT_MERGED", resource_type="contact", resource_id=str(primary_id),
            action_metadata={"merged_from": str(secondary_id)},
        )
        await DashboardService.invalidate_cache(actor.organization_id)
        return await self.contact_repo.get_contact_by_id(actor.organization_id, primary_id)

    # --- Relationships ---
    async def add_relationship(self, actor: User, contact_id: uuid.UUID, related_contact_id: uuid.UUID, relationship_type: str):
        from app.models.contact_relationship import ContactRelationship
        if contact_id == related_contact_id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="A contact cannot relate to itself")
        await self.get_contact(actor, contact_id)  # scoping
        await self.get_contact(actor, related_contact_id)  # both must exist in org
        rel = ContactRelationship(
            organization_id=actor.organization_id, contact_id=contact_id,
            related_contact_id=related_contact_id, relationship_type=relationship_type, created_by=actor.id,
        )
        self.db.add(rel)
        await self.db.flush()
        await self.db.refresh(rel)
        related = await self.contact_repo.get_contact_by_id(actor.organization_id, related_contact_id)
        return {
            "id": rel.id, "contact_id": rel.contact_id, "related_contact_id": rel.related_contact_id,
            "relationship_type": rel.relationship_type,
            "related_contact_name": f"{related.first_name} {related.last_name}".strip() if related else None,
        }

    async def list_relationships(self, actor: User, contact_id: uuid.UUID) -> list[dict]:
        from app.models.contact_relationship import ContactRelationship
        await self.get_contact(actor, contact_id)
        res = await self.db.execute(
            select(ContactRelationship).filter(
                ContactRelationship.contact_id == contact_id,
                ContactRelationship.is_deleted == False,
            )
        )
        rels = list(res.scalars().all())
        related_ids = [r.related_contact_id for r in rels]
        names = {}
        if related_ids:
            cres = await self.db.execute(select(Contact.id, Contact.first_name, Contact.last_name).filter(Contact.id.in_(related_ids)))
            for cid, fn, ln in cres.all():
                names[cid] = f"{fn} {ln}".strip()
        return [
            {"id": r.id, "contact_id": r.contact_id, "related_contact_id": r.related_contact_id,
             "relationship_type": r.relationship_type, "related_contact_name": names.get(r.related_contact_id)}
            for r in rels
        ]

    async def delete_relationship(self, actor: User, contact_id: uuid.UUID, relationship_id: uuid.UUID) -> None:
        from app.models.contact_relationship import ContactRelationship
        await self.get_contact(actor, contact_id)
        res = await self.db.execute(
            select(ContactRelationship).filter(
                ContactRelationship.id == relationship_id,
                ContactRelationship.contact_id == contact_id,
                ContactRelationship.organization_id == actor.organization_id,
                ContactRelationship.is_deleted == False,
            )
        )
        rel = res.scalars().first()
        if not rel:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Relationship not found")
        rel.is_deleted = True
        self.db.add(rel)
        await self.db.flush()
